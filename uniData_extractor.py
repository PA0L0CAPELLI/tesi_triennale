from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import TimeoutException
import pandas as pd
import os
from openpyxl import load_workbook

#functions-------------------------------------------------------------

def estrai_insegnamento(driver):
    #selezione la parte interessata
    selector = 'dd+ dt'
    #Attendo che gli elementi siano presenti
    try:
        links = WebDriverWait(driver, 4).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, selector))
        )
    except TimeoutException:
        return {
            "dati_presenti": False,
            "motivo": "Elementi 'dd+ dt' non trovati entro 1s",
            "url": driver.current_url
        }


    # Itera su ogni elemento in links
    for link in links:
        try:
            # Clicca sull'elemento per rivelare il contenuto
            link.click()
            
            
        except Exception as e:
            print(f"Errore durante il clic su un elemento: {e}")

    content_selector = '.accordion'
    contents = WebDriverWait(driver, 4).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, content_selector)))
    
    all_data = []

    title_selector = '.u-filetto'
    titles = WebDriverWait(driver, 4).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, title_selector)))


    # Estrazione dei dati
    for title in titles:
        title = title.text.strip()  
        id_insegnamento = title.split("]")[0].strip("[")     
        titolo_insegnamento = title.split("-")[1]          

    # Creazione del dizionario
    insegnamento = {
        "id_insegnamento": id_insegnamento,
        "titolo_insegnamento": titolo_insegnamento
    }

    all_data.append(insegnamento)


    for dl in contents:
        dt_elements = dl.find_elements(By.TAG_NAME, "dt")
        dd_elements = dl.find_elements(By.TAG_NAME, "dd")

        data = {}
        for dt, dd in zip(dt_elements, dd_elements):
            key = dt.text.strip()
            value = dd.text.strip()
            data[key] = value

        all_data.append(data)
    # for content in contents:
    #    print(content.text)
    # Rimuovi la chiave 'Informazioni generali' se esiste
    data.pop('Informazioni generali', None)

    # Unione dei due dizionari
    insegnamento = {**all_data[0], **all_data[1]}
    return insegnamento
            
#---------------------------------------------------------------------end functions





# Crea l'istanza del driver Chrome usando Selenium Manager 
driver = webdriver.Chrome()

# Apri sito web
driver.get('https://uniud.coursecatalogue.cineca.it')

#accetta i cookies
accept_cookies_selector = 'c-p-bn'
accept_cookies = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.ID, accept_cookies_selector)))
accept_cookies.click()

all_data = []


#seleziono l'anno di immatricolazione

a_immatricolazione_selector = '#offerta-formativa'
select_element = driver.find_element(By.CSS_SELECTOR, a_immatricolazione_selector)
select = Select(select_element) # crea un oggetto Select
selected_option = select.first_selected_option # ottieni l'opzione selezionata

all_data.append({"anno_immatricolazione": selected_option.text})

#entra nei dipartimenti------------------------------------------------------
dipartimenti_selector = '.u-font-text'
dipartimenti = WebDriverWait(driver, 10).until(
    EC.presence_of_all_elements_located((By.CSS_SELECTOR, dipartimenti_selector)))

for i in range(len(dipartimenti)):
    dipartimenti = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, dipartimenti_selector)))
    dipartimenti[i].click()


    name_dipartimento_selector = '.corsi-group-title'
    name_dipartimento = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, name_dipartimento_selector)))
    all_data.append({"dipartimento": name_dipartimento.text.strip()})
    print("Dipartimento: ", name_dipartimento.text.strip())

        # entra nei corsi-----------------------------------------------------
    corsi_selector = '#main-content a'
    corsi = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, corsi_selector))
        )

        #estrai tutti gli href dei link trovati
    links_corsi = []
    for corso in corsi:
            href = corso.get_attribute("href")
            if href:  # solo se l'attributo esiste
                links_corsi.append(href)

        #accedi alle pagine web associate ai link
    for link in links_corsi:
            driver.get(link)  # naviga alla pagina del corso

            #ESTRAZIONE DATI CORSO
            #estrai il nome del corso
            name_corso_selector = '.u-filetto'
            name_corso = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, name_corso_selector))
            )
            name_corso = name_corso.text.strip()  
            id_corso = name_corso.split("]")[0].strip("[")     
            name_corso = name_corso.split("]")[1]       
            corso = {
                "id_corso": id_corso,
                "titolo_corso": name_corso

            }

            #prendo il contenitore delle informazioni
            contenitore = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".accordion"))
            )
            
            # =========================
            # Sezione: Informazioni generali
            # =========================
            try:
                # Trova il <dt> della sezione il cui <div> interno contiene il testo "Informazioni generali"
                dt_info = contenitore.find_element(
                    By.XPATH,
                    ".//dt[.//div[contains(normalize-space(.), 'Informazioni generali')]]"
                )
                # Prende il <dd> immediatamente successivo: è il contenuto della sezione
                dd_info = dt_info.find_element(By.XPATH, "following-sibling::dd[1]")

                # Dentro il <dd> ci sono tanti sotto-blocchi <dl>, ognuno con una coppia <dt>(chiave) / <dd>(valore)
                dls = dd_info.find_elements(By.XPATH, ".//dl[.//dt]")
                for dl in dls:
                    # Estrae la chiave (testo del <dt>)
                    key = dl.find_element(By.TAG_NAME, "dt").text.strip()
                    # Prende il <dd> associato (il valore)
                    dd = dl.find_element(By.TAG_NAME, "dd")

                    # Inizializza contenuto testuale e (se presente) URL
                    text = ""
                    url = None
                    try:
                        # Se dentro il <dd> c'è un link <a>, usa il testo del link e cattura anche l'href
                        a = dd.find_element(By.TAG_NAME, "a")
                        text = a.text.strip()
                        url = a.get_attribute("href")
                    except:
                        try:
                            # Altrimenti, se c'è uno <span>, usa il testo dello span
                            text = dd.find_element(By.TAG_NAME, "span").text.strip()
                        except:
                            # Fallback: prendi tutto il testo del <dd>
                            text = dd.text.strip()

                    # Normalizza la chiave per evitare ritorni a capo/spazi strani
                    key_norm = key.replace("\n", " ").replace("\r", " ").strip()

                    # Salva nel dizionario con un prefisso per non confondere con altre sezioni
                    corso[f"generali_{key_norm}"] = text
                    # Se c'è un URL (es. "Sito web"), salvalo in una colonna dedicata
                    if url:
                        corso[f"generali_{key_norm}_url"] = url
            except:
                # Se la sezione non esiste/è nascosta o cambia struttura, evita che il codice si blocchi
                pass

            # =========================
            # Sezione: Programma, testi e obiettivi
            # =========================
            # default
            corso["programma_disponibile"] = False
            corso["programma_testo"] = "NON TROVATO"

            try:
                # cerca un <dt> che contenga "programma" (case-insensitive)
                dt = next(iter(contenitore.find_elements(
                    By.XPATH,
                    ".//dt[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'programma')]"
                )), None)

                if dt:
                    try: dt.click()
                    except: driver.execute_script("arguments[0].click();", dt)

                    dd = next(iter(dt.find_elements(By.XPATH, "following-sibling::dd[1]")), None)
                    txt = (dd.get_attribute("innerText") or "").strip() if dd else ""
                    if txt:
                        corso["programma_disponibile"] = True
                        corso["programma_testo"] = txt
            except Exception:
                pass  # resta coi default
            
            
            all_data.append(corso)

            #entro nella pagina degli insegnamenti
            btn_insegnamenti=WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, ".active~ li+ li a"))
            )
            btn_insegnamenti.click()
            #prendo i link degli insegnamenti
            insegnamenti_selector = '.flex-container a'
            insegnamenti = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, insegnamenti_selector))
            )

            #estrai tutti gli href dei link trovati
            links_insegnamenti = []
            for insegnamento in insegnamenti:
                href = insegnamento.get_attribute("href")
                if href:  # solo se l'attributo esiste
                    links_insegnamenti.append(href)
            
            #entro negli insegnamenti -------------------------------------------------
            for link in links_insegnamenti:
                driver.get(link)  # naviga alla pagina dell'insegnmento
                #ESTRAZIONE DATI INSEGNAMENTO

                #controllo che ci sia un insegnamento diviso in moduli
                try:
                    xpath = ("//div[contains(@class,'insegnamento-links')]""//p[contains(translate(.,'ABCDEFGHIJKLMNOPQRSTUVWXYZÀÈÉÌÒÙ','abcdefghijklmnopqrstuvwxyzàèéìòù'),'diviso in moduli')]")
                    wait = WebDriverWait(driver, 4)
                    p = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
                    div_moduli = driver.find_elements(By.CSS_SELECTOR, "div.insegnamento-links li a")
                    links_moduli = []
                    if div_moduli:
                        for modulo in div_moduli:
                            href = modulo.get_attribute("href")
                            if href:  # solo se l'attributo esiste
                                links_moduli.append(href)
                    
                        
                    for link_modulo in links_moduli:
                        driver.get(link_modulo)  # naviga alla pagina del modulo
                        insegnamento = estrai_insegnamento(driver)
                        all_data.append(insegnamento)
                        #torna alla pagina dell'insegnamento
                        driver.back()
                except:
                    insegnamento = estrai_insegnamento(driver)
                    all_data.append(insegnamento)
                #torna alla pagina del corso
                driver.back()


            
            
            #torna al blocco INFO
            driver.back()
            # torna indietro alla pagina della lista corsi
            driver.back()
        #torna indietro ai dipartimenti
    driver.back()

# Creazione del DataFrame dai nuovi dati
df_nuovi = pd.DataFrame(all_data)

excel_path = "uniud_dati.xlsx"

# Se il file non esiste, lo crea con header
if not os.path.exists(excel_path):
    df_nuovi.to_excel(excel_path, index=False, sheet_name="Sheet1")
else:
    # Se esiste, aggiunge i nuovi dati sotto senza riscrivere header
    wb = load_workbook(excel_path)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    startrow = ws.max_row if ws.max_row is not None else 0

    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        writer.book = wb
        writer.sheets = {w.title: w for w in wb.worksheets}
        df_nuovi.to_excel(
            writer,
            sheet_name=ws.title,
            index=False,
            header=False,       # non riscrive le etichette
            startrow=startrow
        )
    wb.save(excel_path)
dipartimenti = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, dipartimenti_selector)))
driver.back()
    

# Chiudi il driver
driver.quit()








